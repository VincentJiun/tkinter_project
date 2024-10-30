from openpyxl import Workbook, load_workbook
import os

class Excel():
    def __init__(self, path):
        self.path = f'./data/{path}'
        if os.path.exists(self.path):
            self.wb = load_workbook(self.path)
        else:
            self.wb = Workbook()
            self.wb.save(self.path)  # 只在創建新工作簿時保存

class ExcelCMS(Excel):
    def __init__(self, path):
        super().__init__(path)

    def create(self, *args):
        self.ws = self.wb.active
        self.ws.append(args)
        self.wb.save(self.path)  # 保存檔案
        # self.wb.close()  # 不要在這裡關閉，保留後續使用

    def read_all_datas(self):
        data = []
        self.ws = self.wb.active
        for row in self.ws.iter_rows(values_only=True, min_row=2):
            # 將每一行的資料轉換為列表，並添加到 data 中
            data.append(list(row))
        # 關閉工作簿
        self.wb.close()

        return data
    
    def get_column_data(self, column_letter):
        column_data = []
        self.ws = self.wb.active
        for cell in self.ws[column_letter]:
            column_data.append(cell.value)  # 獲取每個單元格的值

        return column_data[1:] # 扣除第一個值(標題)
    
    def get_row_data(self, row_index):
        row_data = [cell.value for cell in self.ws[row_index+2]]
        return row_data

    
    def update_row(self, row_index, updated_data):
        # 打開 Excel 並更新對應行
        self.ws = self.wb.active
        for col_num, value in enumerate(updated_data, start=1):
            self.ws.cell(row=row_index + 2, column=col_num).value = value  # row_index+2是因為第一行是標題
        self.wb.save(self.path)

    def delete_row(self, row_index):
        # 打開 Excel 並更新對應行
        self.ws = self.wb.active
        self.ws.delete_rows(row_index + 2)
        self.wb.save(self.path)

class ExcelQuote(Excel):
    def __init__(self, path):
        self.path = f'./quote/{path}.xlsx'
        if os.path.exists('./template/repair.xlsx'):
            self.wb = load_workbook('./template/repair.xlsx')
            self.wb.save(self.path)
        else:
            pass

    def modify_quote(self, q_time, custom, *list):
        self.wb_quote = load_workbook(self.path)
        self.ws = self.wb_quote.active
        # 將 data 填入 Excel 表格
        for col, column_data in enumerate(list, start=2):  # 逐列處理資料 =SUM(D9*F9)
            for row, value in enumerate(column_data, start=9):
                self.ws.cell(row=row, column=col, value=value)

        for row, value in enumerate(list[1], start=1):
            self.ws.cell(row=row+8, column=1, value=str(row))
            formula = f"=SUM(D{row+8}*F{row+8})"
            self.ws.cell(row=row+8, column=7, value=formula)

            # for col in [4, 6]:  # D列和F列
            #     cell = self.ws.cell(row=row+8, column=col)
            #     cell.number_format = '0'  # 數字格式

        self.ws['A6'] = f'客戶名稱：{custom}' # 微軟正黑體 Light, 14 ,bold
        str_time = q_time.split('/')
        self.ws['F6'] = f'報價日期：{str_time[0]}年{str_time[1]}月{str_time[2]}日'

        self.wb_quote.save(self.path)
        # self.wb_quote.close()
        
